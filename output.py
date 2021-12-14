# includes expansions of the original code
# getting more data - e.g. closing + opening + % change of the stock
# this is the code between "stocks" and "output"


#### IMPORTS #####
 
from selenium import webdriver
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
import pandas as pd
from datetime import date
from openpyxl import load_workbook 

from openpyxl.chart import LineChart, Reference


#### CODE ####

df = pd.read_excel("stocks.xls")

# this time data is a list of lists
data = []

service = Service("/Users/rebec/Downloads/chromedriver")
service.start()

for stock in df["stocks"]:
    info = []
    
    driver = webdriver.Remote(service.service_url)
    yahoo = driver.get("https://au.finance.yahoo.com/")
    search = driver.find_element(By.ID, 'yfin-usr-qry')
    search.send_keys("{0}".format(stock))
    time.sleep(5)
    pg = search.send_keys(Keys.RETURN)
    time.sleep(7)
    
    prev_close = driver.find_element(By.XPATH, "//*[@id='quote-summary']/div[1]/table/tbody/tr[1]/td[2]").get_attribute("innerHTML")
    op = driver.find_element(By.XPATH, "//*[@id='quote-summary']/div[1]/table/tbody/tr[2]/td[2]").get_attribute("innerHTML")
    change = driver.find_element(By.XPATH, "//*[@id='quote-header-info']/div[3]/div[1]/div/fin-streamer[3]/span").get_attribute("innerHTML").strip("()")
    
    info.append(date.today())
    info.append(stock)
    info.append(float(prev_close))
    info.append(float(op))
    info.append(change)
    
    data.append(info)
    
    driver.close()
    
driver.quit()

#writing to excel

#pre-existing excel book
name = 'output.xlsx'
wb = load_workbook(name)
page = wb.active

# appending new row with the list of data values for each stock
for stock in data:
    page.append(stock)

# create 4 worksheets for each stock with headings in each
if len(wb.sheetnames) != len(df["stocks"]) + 1:
    for stock in df["stocks"]:
        ws = wb.create_sheet(stock)

        # Setting the headings of each column
        ws['A1'] = 'date'
        ws['B1'] = 'stock_id'
        ws['C1'] = 'prev_close'
        ws['D1'] = 'prev_open' 
        ws['E1'] = 'change'

        # Setting the dimensions of each column
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15


# appending new row with the list of data values for each stock
n = 1
sheets = wb.sheetnames
zip_data = zip(wb, data)
for sheet, stock in zip_data:
    ws = wb[sheets[n]]
    ws.append(stock) 
    n += 1

wb.save(filename=name)

# Creating line charts
sheets = wb.sheetnames
for sheet in sheets:
    ws = wb[sheets[n]]
    values = Reference(ws, min_col = 1, max_col = 4)
    chart = LineChart()
  
    chart.add_data(values)
    chart.title = f" {sheet.name} "
    chart.x_axis.title = " Dates "
    chart.y_axis.title = " Amount "

    sheet.add_chart(chart, "F2")

wb.save(filename=name)

